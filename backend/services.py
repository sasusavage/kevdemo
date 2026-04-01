"""
Service layer — business logic, database transactions, analytics,
forecasting, audit logging, and distributor intelligence.

Routes delegate here; this keeps controllers thin.
"""
from datetime import datetime, timezone, timedelta
from sqlalchemy import func, extract, and_, case
from extensions import db
from models import Product, Distributor, Sale, StockTransaction


# ══════════════════════════════════════════════
#  INVENTORY
# ══════════════════════════════════════════════

def get_all_inventory():
    """Return every product with computed stock status."""
    products = Product.query.order_by(Product.name).all()
    return [p.to_dict() for p in products]


def get_inventory_summary():
    """Aggregate inventory KPIs for the dashboard cards."""
    total_sku = db.session.query(func.count(Product.id)).scalar() or 0
    low_stock = (
        db.session.query(func.count(Product.id))
        .filter(Product.quantity > 0, Product.quantity <= Product.min_stock_level)
        .scalar()
    ) or 0
    critical = (
        db.session.query(func.count(Product.id))
        .filter(
            Product.quantity > 0,
            Product.quantity <= (Product.min_stock_level / 2),
        )
        .scalar()
    ) or 0
    out_of_stock = (
        db.session.query(func.count(Product.id))
        .filter(Product.quantity <= 0)
        .scalar()
    ) or 0
    inventory_value = (
        db.session.query(
            func.sum(Product.quantity * Product.price)
        ).scalar()
    ) or 0

    return {
        "total_sku": total_sku,
        "low_stock_alerts": low_stock,
        "critical_alerts": critical,
        "out_of_stock": out_of_stock,
        "inventory_value": float(inventory_value),
    }


# ══════════════════════════════════════════════
#  SALES (transactional + audit logging)
# ══════════════════════════════════════════════

def record_sale(product_id: int, distributor_id: int, quantity_sold: int,
                sale_date: str | None = None):
    """
    Record a sale inside an atomic database transaction.

    1. Lock the product row (SELECT ... FOR UPDATE) to prevent races.
    2. Verify sufficient stock.
    3. Deduct quantity.
    4. Insert sale record.
    5. Log a StockTransaction audit entry.
    6. Commit — or rollback on any error.
    """
    product = (
        db.session.query(Product)
        .filter_by(id=product_id)
        .with_for_update()
        .first()
    )
    if product is None:
        raise ValueError(f"Product with id {product_id} not found.")

    distributor = Distributor.query.get(distributor_id)
    if distributor is None:
        raise ValueError(f"Distributor with id {distributor_id} not found.")

    if quantity_sold <= 0:
        raise ValueError("Quantity sold must be a positive integer.")

    if product.quantity < quantity_sold:
        raise ValueError(
            f"Insufficient stock. Available: {product.quantity}, "
            f"requested: {quantity_sold}."
        )

    # --- begin atomic mutation ---
    qty_before = product.quantity
    product.quantity -= quantity_sold
    qty_after = product.quantity

    # Step 3: Record Sale — snapshot cost, profit, and commission at time of sale
    total_price = product.price * quantity_sold
    unit_cost = product.cost_price
    profit = total_price - (unit_cost * quantity_sold)
    commission_earned = float(total_price) * float(distributor.commission_rate)

    sale_date_dt = (
        datetime.fromisoformat(sale_date)
        if sale_date
        else datetime.now(timezone.utc)
    )

    sale = Sale(
        product_id=product_id,
        distributor_id=distributor_id,
        quantity_sold=quantity_sold,
        total_price=total_price,
        unit_cost=unit_cost,
        profit=profit,
        commission_earned=commission_earned,
        timestamp=sale_date_dt,
    )
    db.session.add(sale)
    db.session.flush()  # get sale.id

    # Audit log
    txn = StockTransaction(
        product_id=product_id,
        transaction_type="SALE",
        quantity_change=-quantity_sold,
        quantity_before=qty_before,
        quantity_after=qty_after,
        reason=f"Sale to {distributor.name}",
        reference_id=sale.id,
        timestamp=sale_date_dt,
    )
    db.session.add(txn)
    db.session.commit()

    return sale.to_dict()


def record_restock(product_id: int, quantity: int, reason: str = "Manual restock"):
    """
    Add stock to a product and log the transaction.
    """
    product = (
        db.session.query(Product)
        .filter_by(id=product_id)
        .with_for_update()
        .first()
    )
    if product is None:
        raise ValueError(f"Product with id {product_id} not found.")
    if quantity <= 0:
        raise ValueError("Restock quantity must be positive.")

    qty_before = product.quantity
    product.quantity += quantity
    qty_after = product.quantity

    txn = StockTransaction(
        product_id=product_id,
        transaction_type="RESTOCK",
        quantity_change=quantity,
        quantity_before=qty_before,
        quantity_after=qty_after,
        reason=reason,
    )
    db.session.add(txn)
    db.session.commit()

    return product.to_dict()


def record_adjustment(product_id: int, new_quantity: int, reason: str = "Manual adjustment"):
    """
    Set product stock to an exact number (inventory correction).
    """
    product = (
        db.session.query(Product)
        .filter_by(id=product_id)
        .with_for_update()
        .first()
    )
    if product is None:
        raise ValueError(f"Product with id {product_id} not found.")

    qty_before = product.quantity
    change = new_quantity - qty_before
    product.quantity = new_quantity

    txn = StockTransaction(
        product_id=product_id,
        transaction_type="ADJUSTMENT",
        quantity_change=change,
        quantity_before=qty_before,
        quantity_after=new_quantity,
        reason=reason,
    )
    db.session.add(txn)
    db.session.commit()

    return product.to_dict()


def get_stock_transactions(product_id: int | None = None, limit: int = 50):
    """Return recent stock audit log entries."""
    q = StockTransaction.query.order_by(StockTransaction.timestamp.desc())
    if product_id:
        q = q.filter(StockTransaction.product_id == product_id)
    return [t.to_dict() for t in q.limit(limit).all()]


# ══════════════════════════════════════════════
#  ANALYTICS & FORECASTING
# ══════════════════════════════════════════════

def create_product(data: dict):
    """Register a new SKU in the database."""
    name = data.get("name")
    sku = data.get("sku")
    price = data.get("price", 0.0)
    cost_price = data.get("cost_price", 0.0) # Added cost_price
    quantity = data.get("quantity", 0)
    category = data.get("category", "General")
    min_stock = data.get("min_stock_level", 20)
    
    if not name or not sku:
        raise ValueError("Product Name and SKU are required.")
    
    # Check if SKU already exists
    existing = Product.query.filter_by(sku=sku).first()
    if existing:
        raise ValueError(f"SKU '{sku}' is already registered.")
        
    p = Product(
        name=name,
        sku=sku,
        price=price,
        cost_price=cost_price, # Added cost_price
        quantity=quantity,
        category=category,
        min_stock_level=min_stock
    )
    db.session.add(p)
    
    # Log initial inventory as a 'RESTOCK' if qty > 0
    if quantity > 0:
        db.session.flush()
        txn = StockTransaction(
            product_id=p.id,
            transaction_type="RESTOCK",
            quantity_change=quantity,
            quantity_before=0,
            quantity_after=quantity,
            reason="Initial stock on registration"
        )
        db.session.add(txn)
        
    db.session.commit()
    return p.to_dict()


def export_to_excel():
    """Generate a multi-sheet Excel report of the entire system."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Inventory ---
    ws1 = wb.active
    ws1.title = "Current Inventory"
    headers = ["ID", "Name", "SKU", "Quantity", "Price (GHC)", "Category", "Min Stock", "Status"]
    ws1.append(headers)
    
    products = Product.query.order_by(Product.name).all()
    for p in products:
        ws1.append([p.id, p.name, p.sku, p.quantity, float(p.price), p.category, p.min_stock_level, p.stock_status])
        
    # --- Sheet 2: Recent Sales ---
    ws2 = wb.create_sheet("Sales Ledger")
    ws2.append(["ID", "Timestamp", "Product", "Distributor", "Qty Sold", "Total (GHC)"])
    sales = Sale.query.order_by(Sale.timestamp.desc()).limit(1000).all()
    for s in sales:
        ws2.append([s.id, s.timestamp.strftime("%Y-%m-%d %H:%M"), s.product.name, s.distributor.name, s.quantity_sold, float(s.total_price)])
        
    # Formatting
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def get_forecast():
    """
    For each product, calculate:
    - Average daily burn rate (last 30 days of sales)
    - Estimated days until out of stock
    - Stock alert level
    """
    now = datetime.now(timezone.utc)
    thirty_days_ago = now - timedelta(days=30)

    # Get last-30-day sales volume per product
    burn_data = (
        db.session.query(
            Sale.product_id,
            func.coalesce(func.sum(Sale.quantity_sold), 0).label("total_sold"),
        )
        .filter(Sale.timestamp >= thirty_days_ago)
        .group_by(Sale.product_id)
        .all()
    )
    burn_map = {row.product_id: int(row.total_sold) for row in burn_data}

    products = Product.query.order_by(Product.name).all()
    forecasts = []

    for p in products:
        sold_30d = burn_map.get(p.id, 0)
        daily_burn = round(sold_30d / 30, 2)

        if daily_burn > 0 and p.quantity > 0:
            days_remaining = round(p.quantity / daily_burn, 1)
        elif p.quantity <= 0:
            days_remaining = 0
        else:
            days_remaining = None  # no recent sales, can't predict

        forecasts.append({
            "product_id": p.id,
            "product_name": p.name,
            "sku": p.sku,
            "current_stock": p.quantity,
            "min_stock_level": p.min_stock_level,
            "sold_last_30_days": sold_30d,
            "avg_daily_burn_rate": daily_burn,
            "estimated_days_until_oos": days_remaining,
            "status": p.stock_status,
            "alert_level": p.alert_level,
            "restock_recommended": (
                daily_burn > 0 and p.quantity <= p.min_stock_level
            ),
        })

    # Sort: most urgent first (lowest days remaining)
    forecasts.sort(
        key=lambda x: (
            x["estimated_days_until_oos"]
            if x["estimated_days_until_oos"] is not None
            else 9999
        )
    )
    return forecasts


# ══════════════════════════════════════════════
#  AUTOMATED ALERTS & DAILY SUMMARY
# ══════════════════════════════════════════════

def get_stock_alerts():
    """Return products that are below their min_stock_level threshold."""
    products = (
        Product.query
        .filter(Product.quantity <= Product.min_stock_level)
        .order_by(Product.quantity.asc())
        .all()
    )
    return [p.to_dict() for p in products]


def get_daily_summary():
    """Returns today's revenue, volume, and top products."""
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = now - timedelta(days=30)

    top_products = (
        db.session.query(
            Product.name,
            db.func.sum(Sale.quantity_sold).label("volume"),
            db.func.sum(Sale.total_price).label("revenue"),
        )
        .join(Sale)
        .filter(Sale.timestamp >= today_start)
        .group_by(Product.name)
        .order_by(db.func.sum(Sale.total_price).desc())
        .limit(3).all()
    )

    today_revenue = db.session.query(db.func.sum(Sale.total_price)).filter(Sale.timestamp >= today_start).scalar() or 0
    today_volume = db.session.query(db.func.sum(Sale.quantity_sold)).filter(Sale.timestamp >= today_start).scalar() or 0
    alert_count = Product.query.filter(Product.quantity <= Product.min_stock_level).count()

    return {
        "date": now.strftime("%Y-%m-%d"),
        "top_3_products": [{"name":r.name, "volume":int(r.volume), "revenue":float(r.revenue)} for r in top_products],
        "today_revenue": float(today_revenue),
        "today_volume": int(today_volume),
        "stock_alerts": alert_count,
    }


def get_distributor_performance():
    """Aggregated metrics for each distributor including profit share."""
    results = (
        db.session.query(
            Distributor.id,
            Distributor.name,
            Distributor.region,
            Distributor.tier,
            Distributor.commission_rate,
            db.func.coalesce(db.func.sum(Sale.quantity_sold), 0).label("total_volume"),
            db.func.coalesce(db.func.sum(Sale.total_price), 0).label("total_revenue"),
            db.func.coalesce(db.func.sum(Sale.profit), 0).label("total_profit"),
            db.func.coalesce(db.func.sum(Sale.commission_earned), 0).label("total_commissions"),
            db.func.count(Sale.id).label("order_count"),
        )
        .outerjoin(Sale)
        .group_by(Distributor.id, Distributor.name, Distributor.region, Distributor.tier, Distributor.commission_rate)
        .order_by(db.func.sum(Sale.total_price).desc().nullslast())
        .all()
    )

    performance = []
    for r in results:
        rev = float(r.total_revenue)
        prof = float(r.total_profit)
        oc = int(r.order_count)
        growth = _distributor_growth(r.id)

        performance.append({
            "id": r.id,
            "name": r.name,
            "region": r.region,
            "tier": _calculate_tier(rev),
            "commission_rate": float(r.commission_rate),
            "ref_code": f"D{r.id}",
            "total_volume": int(r.total_volume),
            "total_revenue": rev,
            "total_profit": prof,
            "total_commissions": float(r.total_commissions),
            "order_count": oc,
            "avg_order_value": round(rev / oc, 2) if oc > 0 else 0,
            "growth_percent": growth,
        })
    return performance


def _calculate_tier(monthly_revenue: float) -> str:
    """
    Auto-rank distributors into tiers based on monthly revenue.
      Gold:   >= 5000 GHC/month
      Silver: >= 2000 GHC/month
      Bronze: < 2000 GHC/month
    """
    if monthly_revenue >= 5000:
        return "Gold"
    elif monthly_revenue >= 2000:
        return "Silver"
    else:
        return "Bronze"


def _distributor_growth(distributor_id: int) -> float:
    """
    Month-over-month growth for a single distributor.
    Compares current month revenue to previous month revenue.
    """
    now = datetime.now(timezone.utc)
    current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    prev_month_start = (current_month_start - timedelta(days=1)).replace(day=1)

    current_rev = (
        db.session.query(func.coalesce(func.sum(Sale.total_price), 0))
        .filter(
            Sale.distributor_id == distributor_id,
            Sale.timestamp >= current_month_start,
        )
        .scalar()
    )
    prev_rev = (
        db.session.query(func.coalesce(func.sum(Sale.total_price), 0))
        .filter(
            Sale.distributor_id == distributor_id,
            Sale.timestamp >= prev_month_start,
            Sale.timestamp < current_month_start,
        )
        .scalar()
    )

    return calculate_growth_trend(float(current_rev), float(prev_rev))


# ══════════════════════════════════════════════
#  STATS & GROWTH TRENDS
# ══════════════════════════════════════════════

def calculate_growth_trend(current_value: float, previous_value: float) -> float:
    """
    Pure function — percentage change between two periods.
    Positive for growth, negative for decline, 0.0 when no baseline.
    """
    if previous_value == 0:
        return 100.0 if current_value > 0 else 0.0
    return round(((current_value - previous_value) / previous_value) * 100, 2)


def get_sales_stats():
    """
    Aggregate sales data for dashboard stat cards and trend charts.
    Returns weekly sales, monthly sales, total revenue, and a 12-month
    revenue timeline with month-over-month growth.
    """
    now = datetime.now(timezone.utc)
    week_start = now - timedelta(days=now.weekday())
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Aggregate KPIs
    weekly_sales = (
        db.session.query(func.coalesce(func.sum(Sale.quantity_sold), 0))
        .filter(Sale.timestamp >= week_start)
        .scalar()
    )
    weekly_revenue = (
        db.session.query(func.coalesce(func.sum(Sale.total_price), 0))
        .filter(Sale.timestamp >= week_start)
        .scalar()
    )
    monthly_sales = (
        db.session.query(func.coalesce(func.sum(Sale.quantity_sold), 0))
        .filter(Sale.timestamp >= month_start)
        .scalar()
    )
    monthly_revenue = (
        db.session.query(func.coalesce(func.sum(Sale.total_price), 0))
        .filter(Sale.timestamp >= month_start)
        .scalar()
    )
    total_revenue = (
        db.session.query(func.coalesce(func.sum(Sale.total_price), 0)).scalar()
    )
    total_sales_volume = (
        db.session.query(func.coalesce(func.sum(Sale.quantity_sold), 0)).scalar()
    )

    # Previous-period comparison
    prev_week_start = week_start - timedelta(days=7)
    prev_weekly_sales = (
        db.session.query(func.coalesce(func.sum(Sale.quantity_sold), 0))
        .filter(Sale.timestamp >= prev_week_start, Sale.timestamp < week_start)
        .scalar()
    )

    prev_month_start = (month_start - timedelta(days=1)).replace(day=1)
    prev_monthly_sales = (
        db.session.query(func.coalesce(func.sum(Sale.quantity_sold), 0))
        .filter(Sale.timestamp >= prev_month_start, Sale.timestamp < month_start)
        .scalar()
    )

    # Monthly revenue timeline (Trailing 12 Months)
    timeline = []
    
    # Generate 12 months of slots ending with current month
    for i in range(11, -1, -1):
        target_date = now - timedelta(days=i * 30) # approx 30-day steps back
        target_month = target_date.month
        target_year = target_date.year
        
        # Get data for this specific year/month
        row = (
            db.session.query(
                func.sum(Sale.total_price).label("revenue"),
                func.sum(Sale.quantity_sold).label("volume"),
            )
            .filter(extract("year", Sale.timestamp) == target_year)
            .filter(extract("month", Sale.timestamp) == target_month)
            .first()
        )
        
        rev = float(row.revenue) if row and row.revenue else 0.0
        vol = int(row.volume) if row and row.volume else 0
        
        # Calculate growth relative to the previous entry in our generated timeline
        prev_rev = timeline[-1]["revenue"] if len(timeline) > 0 else 0.0
        growth = calculate_growth_trend(rev, prev_rev)
        
        timeline.append({
            "year": target_year,
            "month": target_month,
            "revenue": rev,
            "volume": vol,
            "growth_percent": growth,
        })

    return {
        "weekly_sales": int(weekly_sales),
        "weekly_revenue": float(weekly_revenue),
        "weekly_growth": calculate_growth_trend(
            float(weekly_sales), float(prev_weekly_sales)
        ),
        "monthly_sales": int(monthly_sales),
        "monthly_revenue": float(monthly_revenue),
        "monthly_growth": calculate_growth_trend(
            float(monthly_sales), float(prev_monthly_sales)
        ),
        "total_revenue": float(total_revenue),
        "total_sales_volume": int(total_sales_volume),
        "monthly_timeline": timeline,
    }


# ══════════════════════════════════════════════
#  CUSTOM REPORTS
# ══════════════════════════════════════════════

def generate_report(start_date: str | None = None, end_date: str | None = None,
                    category: str | None = None, distributor_id: int | None = None):
    """
    Generate a custom performance report filtered by date range,
    product category, and/or distributor.

    Returns per-product breakdown with aggregated metrics.
    """
    q = (
        db.session.query(
            Product.name.label("product_name"),
            Product.sku,
            Product.category,
            Distributor.name.label("distributor_name"),
            func.sum(Sale.quantity_sold).label("total_volume"),
            func.sum(Sale.total_price).label("total_revenue"),
            func.count(Sale.id).label("order_count"),
            func.min(Sale.timestamp).label("first_sale"),
            func.max(Sale.timestamp).label("last_sale"),
        )
        .join(Sale, Sale.product_id == Product.id)
        .join(Distributor, Sale.distributor_id == Distributor.id)
    )

    # Apply filters
    if start_date:
        q = q.filter(Sale.timestamp >= datetime.fromisoformat(start_date))
    if end_date:
        end_dt = datetime.fromisoformat(end_date).replace(
            hour=23, minute=59, second=59
        )
        q = q.filter(Sale.timestamp <= end_dt)
    if category:
        q = q.filter(Product.category.ilike(f"%{category}%"))
    if distributor_id:
        q = q.filter(Sale.distributor_id == distributor_id)

    rows = (
        q.group_by(
            Product.name, Product.sku, Product.category, Distributor.name
        )
        .order_by(func.sum(Sale.total_price).desc())
        .all()
    )

    total_revenue = sum(float(r.total_revenue) for r in rows)
    total_volume = sum(int(r.total_volume) for r in rows)

    report_data = []
    for r in rows:
        rev = float(r.total_revenue)
        vol = int(r.total_volume)
        orders = int(r.order_count)
        report_data.append({
            "product_name": r.product_name,
            "sku": r.sku,
            "category": r.category,
            "distributor_name": r.distributor_name,
            "total_volume": vol,
            "total_revenue": rev,
            "order_count": orders,
            "avg_order_value": round(rev / orders, 2) if orders > 0 else 0,
            "revenue_share_pct": round((rev / total_revenue) * 100, 2) if total_revenue > 0 else 0,
            "first_sale": r.first_sale.isoformat() if r.first_sale else None,
            "last_sale": r.last_sale.isoformat() if r.last_sale else None,
        })

    return {
        "filters_applied": {
            "start_date": start_date,
            "end_date": end_date,
            "category": category,
            "distributor_id": distributor_id,
        },
        "summary": {
            "total_revenue": total_revenue,
            "total_volume": total_volume,
            "total_line_items": len(report_data),
        },
        "data": report_data,
    }


# ══════════════════════════════════════════════
#  REVENUE GROWTH TRENDS
# ══════════════════════════════════════════════

def get_sales_trends():
    """
    Revenue-focused trend data for the Growth Trends chart.

    Returns:
    - weekly_revenue / monthly_revenue with WoW and MoM % growth
    - daily_breakdown: last 7 days of daily revenue for bar chart
    """
    now = datetime.now(timezone.utc)
    week_ago = now - timedelta(days=7)
    prev_week_ago = now - timedelta(days=14)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    prev_month_start = (month_start - timedelta(days=1)).replace(day=1)

    weekly_revenue = float(
        db.session.query(func.coalesce(func.sum(Sale.total_price), 0))
        .filter(Sale.timestamp >= week_ago)
        .scalar()
    )
    prev_weekly_revenue = float(
        db.session.query(func.coalesce(func.sum(Sale.total_price), 0))
        .filter(Sale.timestamp >= prev_week_ago, Sale.timestamp < week_ago)
        .scalar()
    )
    monthly_revenue = float(
        db.session.query(func.coalesce(func.sum(Sale.total_price), 0))
        .filter(Sale.timestamp >= month_start)
        .scalar()
    )
    prev_monthly_revenue = float(
        db.session.query(func.coalesce(func.sum(Sale.total_price), 0))
        .filter(Sale.timestamp >= prev_month_start, Sale.timestamp < month_start)
        .scalar()
    )

    daily_breakdown = []
    for i in range(6, -1, -1):
        day_start = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        rev = float(
            db.session.query(func.coalesce(func.sum(Sale.total_price), 0))
            .filter(Sale.timestamp >= day_start, Sale.timestamp < day_end)
            .scalar()
        )
        daily_breakdown.append({
            "date": day_start.strftime("%Y-%m-%d"),
            "day": day_start.strftime("%a"),
            "revenue": rev,
        })

    return {
        "weekly_revenue": weekly_revenue,
        "monthly_revenue": monthly_revenue,
        "wow_growth": calculate_growth_trend(weekly_revenue, prev_weekly_revenue),
        "mom_growth": calculate_growth_trend(monthly_revenue, prev_monthly_revenue),
        "daily_breakdown": daily_breakdown,
    }


# ══════════════════════════════════════════════
#  SOCIAL MEDIA QUICK-ORDER WORKFLOW
# ══════════════════════════════════════════════

def record_quick_order(ref: str, sku: str, quantity: int, sale_date: str | None = None):
    """
    Record a sale using a distributor ref code (D{id}) and product SKU.

    Designed for social media DM workflows where the admin receives an order
    via WhatsApp/Instagram and needs to log it in as few steps as possible.

    ref format: "D{distributor_id}" — e.g. "D5" maps to distributor id=5.
    sku: exact product SKU string.
    """
    ref = ref.strip().upper()
    if not ref.startswith("D"):
        raise ValueError(f"Invalid ref '{ref}'. Expected format: D{{id}}, e.g. D5.")
    try:
        distributor_id = int(ref[1:])
    except ValueError:
        raise ValueError(f"Invalid distributor ref: '{ref}'.")

    product = Product.query.filter_by(sku=sku.strip()).first()
    if product is None:
        raise ValueError(f"Product SKU '{sku}' not found.")

    return record_sale(product.id, distributor_id, quantity, sale_date)
